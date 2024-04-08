"""
Openpyxl-based module that copies style elements from a sheet.
"""

from copy import copy
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
import re


class StyleReader:
    def __init__(
        self,
        file,
        sheet=None,
        min_row=0,
        max_row=0,
        min_col=0,
        max_col=0,
        column_widths=[],
    ):
        self.File = None  # Path | File-like object
        self.Data = None  # Workbook
        self.Sheet = None  # String (must be worksheet name)

        self.Orientation = "ByRow"
        self.Orientation_Options = ("ByRow", "ByColumn")

        self.Header_Data = []
        self.Header_MergeCells = []
        self.Header_Dimensions = []
        self.Header_Length = 0

        self.CellStyle_Data = list()
        self.CellStyle_Type = "SampleCell"
        self.CellStyle_Options = (
            "SampleRow",  # Samples single row
            "SampleColumn",  # Samples single column
            "SampleCell",
        )  # Samples single cell

        self.RowRange = (min_row, max_row)
        self.RowDimension = list()  #
        self.RowHidden = list()  #

        self.ColRange = (min_col, max_col)
        self.ColDimension = list()  #
        self.ColHidden = list()

        self.MergedCellRange = (
            list()
        )  # One-dimensional list of tuples (start_row, start_column, end_row, end_column)
        self.Protection = list()  # One-dimensional list
        self.Password = ""
        self.DataValidation = list()  # One-dimensional list
        self.ConditionalFormatting = list()  # One-dimensional list

        self.load(file, sheet, column_widths=column_widths)

    #
    def load(self, file, sheet, column_widths=[]):
        self.File = file
        self.Data = load_workbook(file)

        if sheet == None:
            self.Sheet = self.Data.sheetnames[0]
        else:
            self.Sheet = sheet

        # Row ranges/dimensions
        self.SetRowRange(self.RowRange[0], self.RowRange[1])
        self.GetRowDimensions()

        # Column ranges/dimensions
        self.SetColumnRange(self.ColRange[0], self.ColRange[1])
        self.GetColumnDimensions(column_widths)

        # Other easily getable styles
        self.GetProtection()
        self.GetMergedCells()
        self.GetDataValidation()
        self.GetConditionalFormatting()

    def SetHeaderLength(self, length):
        if length < 0:
            raise ValueError("Header length cannot be less than 0.")
        else:
            self.Header_Length = length

    def GetHeaderStyle(self):
        if self.Header_Length > 0:
            if self.Orientation == "ByRow":
                self.Header_MergeCells = [
                    MergeCellRange
                    for MergeCellRange in self.MergedCellRange
                    if MergeCellRange[1] <= self.Header_Length
                    and MergeCellRange[3] <= self.Header_Length
                ]
                self.Header_Data = [
                    [self.GetCellStyle(cell) for cell in row]
                    for row in self.Data[self.Sheet].iter_rows(
                        max_row=self.Header_Length
                    )
                ]

                if self.RowDimension == []:
                    self.GetRowDimensions()
                self.Header_Dimensions = self.RowDimension[: self.Header_Length]
            elif self.Orientation == "ByColumn":
                self.Header_MergeCells = [
                    MergeCellRange
                    for MergeCellRange in self.MergedCellRange
                    if MergeCellRange[0] <= self.Header_Length
                    and MergeCellRange[2] <= self.Header_Length
                ]

                self.Header_Data = [
                    [self.GetCellStyle(cell) for cell in col]
                    for col in self.Data[self.Sheet].iter_cols(
                        max_col=self.Header_Length
                    )
                ]

    def GetCellStyle(self, cell):
        CellStyle = (
            copy(cell.font),
            copy(cell.fill),
            copy(cell.border),
            copy(cell.alignment),
            copy(cell.protection),
        )

        return CellStyle

    # Samples
    def GetContentStyle(self, row=1, col=1):  # row: 참조 행 번호, col: 참조 열 번호
        if self.CellStyle_Type == "SampleRow":
            self.CellStyle_Data = [
                self.GetCellStyle(cell)
                for cell in list(self.Data[self.Sheet].iter_rows())[row]
            ]
        elif self.CellStyle_Type == "SampleCol":
            self.CellStyle_Data = [
                self.GetCellStyle(cell)
                for cell in list(self.Data[self.Sheet].iter_cols())[col]
            ]
        elif self.CellStyle_Type == "SampleCell":
            self.CellStyle_Data = [self.GetCellStyle(self.Data[self.Sheet][row][col])]

    # Sheet-based style settings
    def GetRowDimensions(self):
        self.RowDimension = [
            self.Data[self.Sheet].row_dimensions[i].height
            for i in range(self.RowRange[0], self.RowRange[1])
        ]

        self.RowHidden = [
            self.Data[self.Sheet].row_dimensions[i].hidden
            for i in range(self.RowRange[0], self.RowRange[1])
        ]

    def GetColumnDimensions(self, column_dimensions):
        if len(column_dimensions) != 0:
            self.ColDimension = column_dimensions
        else:
            self.ColDimension = [
                self.Data[self.Sheet].column_dimensions[get_column_letter(i)].width
                for i in range(self.ColRange[0], self.ColRange[1] + 1)
            ]

        self.ColHidden = [
            self.Data[self.Sheet].column_dimensions[get_column_letter(i)].hidden
            for i in range(self.ColRange[0], self.ColRange[1] + 1)
        ]

    def SetRowRange(self, min_row=0, max_row=0):
        if min_row == 0:
            min_row = self.Data[self.Sheet].min_row
        if max_row == 0:
            max_row = self.Data[self.Sheet].max_row

        self.RowRange = (min_row, max_row)

    def SetColumnRange(self, min_col=0, max_col=0):
        if min_col == 0:
            min_col = self.Data[self.Sheet].min_column
        if max_col == 0:
            max_col = self.Data[self.Sheet].max_column

        self.ColRange = (min_col, max_col)

    def GetMergedCells(self):
        self.MergedCellRange = [
            Range.bounds for Range in self.Data[self.Sheet].merged_cells.ranges
        ]

    def GetProtection(self):
        self.Protection = self.Data[self.Sheet].protection
        self.Password = self.Data[self.Sheet].protection.password

    def GetDataValidation(self):
        self.DataValidation = [
            DV for DV in self.Data[self.Sheet].data_validations.dataValidation
        ]

    def GetConditionalFormatting(self):
        self.ConditionalFormatting = self.Data[self.Sheet].conditional_formatting

    # Functions that apply styles


class StyleWriter:
    def __init__(self, file, style, sheet=None):
        self.Style = style  # StyleReader object
        self.Workbook = None
        self.SheetName = sheet
        self.load(file)

    def load(self, file):
        if type(file) != Workbook:
            self.Workbook = load_workbook(file)

        if self.SheetName == None:
            self.SheetName = self.Workbook.sheetnames[0]

    def ApplyProtection(self):
        self.Workbook[self.SheetName].protection = self.Style.Protection

    def ApplyDataValidation(self, min_row=None, max_row=None):
        for DV in self.Style.DataValidation:
            for ranges in DV.sqref.ranges:
                DV_ranges = re.match(
                    "([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)", ranges.coord
                ).groups()

                if min_row == None:
                    min_row = DV_ranges[1]
                if max_row == None:
                    max_row = DV_ranges[3]

                DV.sqref = f"{DV_ranges[0]}{min_row}:{DV_ranges[2]}{max_row}"
                self.Workbook[self.SheetName].add_data_validation(DV)

        # [self.Workbook[self.SheetName].add_data_validation(DV) for DV in self.Style.DataValidation]

    def ApplyConditionalFormatting(self):
        self.Workbook[self.SheetName].conditional_formatting = (
            self.Style.ConditionalFormatting
        )

    def ApplyColDimensions(self, offset_col=0, min_col=1, max_col=None):
        if max_col == None:
            max_col = len(self.Style.ColDimension) + 1

        for col_no in range(min_col, max_col):
            print(col_no)

            try:
                self.Workbook[self.SheetName].column_dimensions[
                    get_column_letter(col_no + offset_col)
                ].width = self.Style.ColDimension[col_no - 1]
            except IndexError:
                pass

    def ApplyRowDimensions(self, offset_row=0, min_row=1, max_row=None):
        if max_row == None:
            max_row = len(self.Style.RowDimension)

        for row_no in range(min_row, max_row):
            try:
                self.Workbook[self.SheetName].row_dimensions[
                    row_no + offset_row
                ].height = self.Style.RowDimension[row_no - 1]
            except IndexError:
                pass

    def ApplyCellStyle(self, cell, styles):
        cell.font, cell.fill, cell.border, cell.alignment, cell.protection = styles

    def ApplyContentStyles(
        self,
        offset_row=0,
        offset_col=0,
        min_row=None,
        max_row=None,
        min_col=None,
        max_col=None,
        wrap_columns=[],
    ):
        if self.Style.Orientation == "ByRow":
            if min_row == None:
                min_row = self.Style.Header_Length + offset_row + 1
            if max_row == None:
                max_row = self.Style.RowRange[1] + offset_row + 1
            if min_col == None:
                min_col = self.Style.ColRange[0] + offset_col
            if max_col == None:
                max_col = self.Style.ColRange[1] + offset_col

            side = Side(border_style="thin", color="000000")
            thin_border = Border(top=side, bottom=side, left=side, right=side)

            for row in self.Workbook[self.SheetName].iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                try:
                    if self.Style.CellStyle_Type == "SampleCell":
                        for i, cell in enumerate(row):
                            self.ApplyCellStyle(cell, self.Style.CellStyle_Data[0])
                            if get_column_letter(i + 1) in wrap_columns:
                                cell.alignment = Alignment(wrap_text=True)
                            cell.border = thin_border
                    elif self.Style.CellStyle_Type == "SampleRow":
                        map(self.ApplyCellStyle, row, self.Style.CellStyle_Data[0])
                except IndexError:
                    pass
        elif self.Style.Orientation == "ByColumn":
            if min_row == None:
                min_row = self.Style.RowRange[0] + 1
            if max_row == None:
                max_row = self.Style.RowRange[1] + 1
            if min_col == None:
                min_col = self.Style.Header_Length + 1 + offset_col
            if max_col == None:
                max_col = self.Style.ColRange[1] + offset_col

            for col in self.Workbook[self.SheetName].iter_cols(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                try:
                    if self.Style.CellStyle_Type == "SampleCell":
                        for cell in col:
                            self.ApplyCellStyle(cell, self.Style.CellStyle_Data[0])
                    if self.Style.CellStyle_Type == "SampleCol":
                        map(self.ApplyCellStyle, col, self.Style.CellStyle_Data)
                except IndexError:
                    pass

        if self.Style.Orientation == "ByRow":
            self.ApplyColDimensions(offset_col=offset_col)
        elif self.Style.Orientation == "ByColumn":
            self.ApplyRowDimensions(offset_row=offset_row)

    def ApplyHeaderStyles(self, offset_row=0, offset_col=0):
        if self.Style.Orientation == "ByRow":
            for i, width in enumerate(self.Style.ColDimension):
                self.Workbook[self.SheetName].column_dimensions[
                    get_column_letter(i + (1 if offset_col == 0 else offset_col))
                ].width = width
            for row in range(1, self.Style.Header_Length + 1):
                try:
                    self.Workbook[self.SheetName].row_dimensions[row].height = (
                        self.Style.Header_Dimensions[row - 1]
                    )
                except IndexError:
                    print(row)
                for col in range(0, self.Style.ColRange[1] + 1):
                    try:
                        self.ApplyCellStyle(
                            self.Workbook[self.SheetName][row + offset_row][
                                col + offset_col - 1
                            ],
                            self.Style.Header_Data[row - 1][col - 1],
                        )
                    except IndexError:
                        pass
        elif self.Style.Orientation == "ByColumn":
            for row in range(1, self.Style.RowRange[1] + 1):
                for col in range(0, self.Style.Header_Length):
                    self.ApplyCellStyle(
                        self.Workbook[self.SheetName][row + offset_row][
                            col + offset_col
                        ],
                        self.Style.Header_Data[col - 1][row - 1],
                    )

        [
            self.Workbook[self.SheetName].merge_cells(
                start_column=MergeCellRange[0] + offset_col,
                start_row=MergeCellRange[1] + offset_row,
                end_column=MergeCellRange[2] + offset_col,
                end_row=MergeCellRange[3] + offset_row,
            )
            for MergeCellRange in self.Style.Header_MergeCells
        ]

        if self.Style.Orientation == "ByRow":
            self.ApplyRowDimensions(
                offset_row=offset_row, max_row=self.Style.Header_Length
            )
        elif self.Style.Orientation == "ByColumn":
            self.ApplyColDimensions(
                offset_col=offset_col, max_col=self.Style.Header_Length
            )

    def ApplySheetStyle(self, min_row=None, max_row=None):
        self.ApplyDataValidation(min_row=min_row, max_row=max_row)
        self.ApplyConditionalFormatting()
        self.ApplyProtection()

    def Save(self, name):
        self.Workbook.save(name)
