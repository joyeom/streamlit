import streamlit as st
import tracemalloc
import pandas as pd
import Inspection.language as lang

# NAC
from Inspection.NAC import chars_error as nac_chars_error
from Inspection.NAC import emcha_error as nac_emcha_error
from Inspection.NAC import red_error as nac_red_error
from Inspection.NAC import emoji_error as nac_emoji_error

# EXCEL
from Inspection.EXCEL import chars_error as excel_chars_error
from Inspection.EXCEL import emcha_error as excel_emcha_error
from Inspection.EXCEL import red_error as excel_red_error
from Inspection.EXCEL import emoji_error as excel_emoji_error


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment
import pandas as pd

from io import BytesIO
import zipfile

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

st.set_page_config(page_title="Inspection", page_icon="./Inspection/Flitto_symbol.jpg")
st.title("내부검수")

hide_streamlit_style = """
<style>
[data-testid="stToolbar"] {visibility: hidden !important;}
footer {visibility: hidden !important;}
</style>
"""
hide_viewer_badge_style = """
<style>
.viewerBadge_container__r5tak {
    display: none !important;
}
</style>
"""
st.markdown(hide_streamlit_style + hide_viewer_badge_style, unsafe_allow_html=True)


# Base Module
class Widget:
    def __init__(self, root=None):
        self._root = root
        self.initUI()

    def initUI(self):
        pass


# Main Programs
class Main(Widget):
    def __init__(self, root=None):
        self.__DEFAULT_STATE = {
            "uploaded_file": [],
            "environment": "",
            "src_lang": "",
            "tgt_lang": "",
        }
        self.input_validity = False
        self.DEFAULT_SETTINGS()
        super().__init__(root)

    def initUI(self):
        uploaded_file = st.file_uploader(
            "Choose a File", accept_multiple_files=True, type=".xlsx"
        )
        environment = st.selectbox(
            label="작업환경",
            options=("Select an environment", "NAC", "EXCEL"),
            index=0,  # 'Select an environment'가 기본값이 됩니다.
        )

        src_lang_options = ["src_lang"] + lang.bb_language
        tgt_lang_options = ["tgt_lang"] + lang.bb_language
        src_lang = st.selectbox(label="Source lang", options=src_lang_options, index=0)
        tgt_lang = st.selectbox(label="Target lang", options=tgt_lang_options, index=0)

        self.__DEFAULT_STATE["uploaded_file"] = uploaded_file
        self.__DEFAULT_STATE["environment"] = environment
        self.__DEFAULT_STATE["src_lang"] = src_lang
        self.__DEFAULT_STATE["tgt_lang"] = tgt_lang

        if self.check_input_validity():
            for s in self.__DEFAULT_STATE:
                print("after input", s, self.__DEFAULT_STATE[s])

            if st.button("Next", on_click=self.on_next_click):
                pass

    def on_next_click(self):
        processor_menu = ProcessorMenu(self)

    def DEFAULT_SETTINGS(self):
        for state in self.__DEFAULT_STATE:
            if state not in st.session_state:
                st.session_state[state] = self.__DEFAULT_STATE[state]
            print(
                "Default setting st.session_state[state]",
                state,
                st.session_state[state],
                self.__DEFAULT_STATE[state],
            )

    def check_input_validity(self):
        self.input_validity = all(
            [
                self.__DEFAULT_STATE["uploaded_file"] != [],
                self.__DEFAULT_STATE["environment"] != "",
                self.__DEFAULT_STATE["environment"] != "Select an environment",
                self.__DEFAULT_STATE["src_lang"] != "",
                self.__DEFAULT_STATE["src_lang"] != "src_lang",
                self.__DEFAULT_STATE["tgt_lang"] != "",
                self.__DEFAULT_STATE["tgt_lang"] != "tgt_lang",
            ]
        )

        return self.input_validity

    def get_state(self):
        return self.__DEFAULT_STATE


# Processor Menu
class ProcessorMenu(Widget):
    def __init__(self, main_instance, root=None):
        self.state = main_instance.get_state()

        # style
        self.style = {
            "thin_border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "start_idx": (
                4 if self.state["environment"] == "EXCEL" else 2
            ),  # excel은 세번째 row까지 merge되어있기 때문
        }
        super().__init__(root)

    def initUI(self):
        with st.spinner("내부검수 파일 만드는 중.."):
            uploaded_files = self.state["uploaded_file"]
            environment = self.state["environment"]
            src_lang = self.state["src_lang"]
            tgt_lang = self.state["tgt_lang"]
            zip_buffer = self.create_zip_file(
                uploaded_files, src_lang, tgt_lang, environment
            )

            # Zip 파일 다운로드
            st.download_button(
                label="Download Zip",
                data=zip_buffer,
                file_name="Inspection_files.zip",
                mime="application/zip",
            )

    def create_zip_file(self, uploaded_files, src_lang, tgt_lang, environment):
        # BytesIO 객체 생성
        zip_buffer = BytesIO()

        # Zip 파일 생성
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zf:
            for uploaded_file in uploaded_files:
                file_name, error_inspected_df = self.process_data(
                    uploaded_file, src_lang, tgt_lang, environment
                )

                excel_buffer = BytesIO()
                # DataFrame을 Excel 파일로 변환하여 BytesIO에 쓰기
                og_file = BytesIO(uploaded_file.getvalue())
                wb = load_workbook(og_file)
                sheet = wb.active

                # 데이터 처리
                self.write_styled_data_to_sheet(sheet, error_inspected_df, environment)

                # BytesIO 객체에 엑셀 데이터를 쓰기
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                zf.writestr(file_name, excel_buffer.getvalue())

        # BytesIO 객체의 커서 위치를 파일의 시작으로 이동
        zip_buffer.seek(0)

        return zip_buffer

    def process_data(self, f, src_lang, tgt_lang, environment):
        # 파일명 설정
        file_name = f.name

        # DataFrame으로 변환
        df = pd.read_excel(f, engine="openpyxl")

        if environment == "EXCEL":
            df = self.process_excel_data(df, src_lang, tgt_lang)
        elif environment == "NAC":
            df = self.process_nac_data(df, src_lang, tgt_lang)

        return file_name, df

    def process_excel_data(self, df, src_lang, tgt_lang):

        def contains_t2(df):
            if df.iloc[0].str.contains("target").sum() == 2:
                return True

            return False

        only_data = df.iloc[2:,]  # because headers(row 1-3) are merged
        has_t2 = contains_t2(df)

        only_data["Duplicated"] = only_data["f_id"].duplicated(keep=False)
        only_data = excel_red_error.check_redacted(only_data, has_t2)
        only_data = excel_emoji_error.get_emojis(only_data, has_t2)
        only_data = excel_chars_error.get_chars_error(only_data, has_t2)
        only_data = excel_emcha_error.get_emcha(only_data, has_t2, src_lang, tgt_lang)

        error_checked_loc = only_data.columns.get_loc("Duplicated")
        error_checked_df = only_data.iloc[:, error_checked_loc:]

        return error_checked_df

    def process_nac_data(self, df, src_lang, tgt_lang):

        df["Duplicated"] = df["SID"].duplicated(keep=False)
        df = nac_red_error.check_redacted(df)
        df = nac_emoji_error.get_emojis(df)
        df = nac_chars_error.get_chars_error(df)
        df = nac_emcha_error.get_emcha(df, src_lang, tgt_lang)

        error_checked_loc = df.columns.get_loc("Duplicated")
        error_checked_df = df.iloc[:, error_checked_loc:]

        return error_checked_df

    # add "inspected columns" to right of the original file with styles
    def write_styled_data_to_sheet(self, sheet, df, environment):

        max_col = sheet.max_column

        # print(f"self.style in write_styled_data_to_sheet {self.style}")

        for idx, cur_col in enumerate(df.columns):
            cur_letter = get_column_letter(max_col + idx + 1)

            self.style_header_data(sheet, self.style, cur_letter, cur_col)
            self.style_non_header_data(
                df[cur_col], sheet, self.style, cur_letter, cur_col
            )

    def style_header_data(self, sheet, style, col_letter, col):
        if self.state["environment"] == "EXCEL":  # if environment is EXCEL
            # merge header
            sheet.merge_cells(f"{col_letter}1:{col_letter}3")

            # create border for merged cell
            for idx in range(1, 4):
                sheet[f"{col_letter}{idx}"].border = style["thin_border"]
        else:
            sheet[f"{col_letter}1"].border = style["thin_border"]

        # Apply style to the header
        sheet[f"{col_letter}1"].alignment = style["alignment"]
        sheet[f"{col_letter}1"] = col
        sheet.column_dimensions[f"{col_letter}"].width = 20

        fill_color = self.get_fill_color(col)
        sheet[f"{col_letter}1"].fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )  # header

    def style_non_header_data(self, data, sheet, style, col_letter, col):
        for idx, value in enumerate(data, start=style["start_idx"]):
            # print("value ",value,type(value))
            if isinstance(value, (list, str, bool)):
                fill_color = self.get_fill_color(col)
                sheet[f"{col_letter}{idx}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
                sheet[f"{col_letter}{idx}"] = str(value)
                sheet[f"{col_letter}{idx}"].border = style["thin_border"]

            else:
                sheet[f"{col_letter}{idx}"] = value
                sheet[f"{col_letter}{idx}"].border = style["thin_border"]
                pass

    # Apply different color to each error type
    def get_fill_color(self, col):
        if "Duplicated" in col:
            return "FFC0CB"  # pink
        elif "red" in col.lower():
            return "FFFF00"  # yellow
        elif "emoji" in col.lower():
            return "90EE90"  # light green
        elif "chars" in col.lower():
            return "ADD8E6"  # light blue
        elif "emcha" in col.lower():
            return "00FFFF"  # cyan


if __name__ == "__main__":
    Main()
