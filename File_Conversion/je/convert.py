# we only need annotations from json file
# convert json to excel in NAC so that we can do internal audit tool
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from io import BytesIO
import json
import streamlit as st


# Define styles
grey_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
light_green_fill = PatternFill(
    start_color="90EE90", end_color="90EE90", fill_type="solid"
)
light_yellow_fill = PatternFill(
    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
)
light_pink_fill = PatternFill(
    start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
)
light_blue_fill = PatternFill(
    start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
)


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
domains_col = ["ecommerce", "refEdu", "socialMedia", "literature", "other"]
errorType_col = [
    "wrongLang",
    "garbled",
    "nonsensical",
    "noncoherent",
    "typos",
    "missingContext",
    "noneofAbove",
    "freeText",
]


def create_df(data):
    # get "annotation" field in json file
    df = pd.DataFrame(data["annotations"])
    domains = pd.DataFrame([item["domains"] for item in data["annotations"]])
    checkBoxes = pd.DataFrame([item["checkBoxes"] for item in data["annotations"]])
    df = df.drop(columns=["domains", "checkBoxes"])
    df = pd.concat([df, domains, checkBoxes], axis=1)

    return df


def convert_json_to_excel(json_file):
    try:
        data = json.loads(json_file)

    except json.JSONDecodeError:
        st.error("Browses files 안에 보이는 파일들로 선택해주세요")

    annotations = create_df(data)

    # Excel 파일로 변환
    excel_buffer = BytesIO()
    wb = Workbook()  # Create a new Workbook
    sheet = wb.active
    # apply style
    sheet = apply_style(sheet, annotations)
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def apply_header_style(cell, column, header):

    cell.alignment = Alignment(horizontal="center", vertical="center")

    if "relevant" in header or header in ["payload", "translation"]:
        cell.fill = grey_fill if "relevant" in header else light_green_fill
        column.width = 60
    elif header in domains_col:
        cell.fill = light_yellow_fill
    elif header in errorType_col:
        cell.fill = light_pink_fill
    else:  # id, annotatorID, source
        cell.fill = light_blue_fill
        column.width = 10


def write_header_value(cell, header):
    cell.value = header


def apply_cell_style(cell):
    cell.border = thin_border
    cell.alignment = Alignment(vertical="top", wrapText=True)


def write_cell_value(cell, value):
    cell.value = value


def apply_style(sheet, df):
    headers = df.columns

    # header
    for col_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_index)
        column = sheet.column_dimensions[column_letter]
        cell = sheet.cell(row=1, column=col_index)
        write_header_value(cell, header)
        apply_header_style(cell, column, header)

    # non-header
    for col_index, col_name in enumerate(headers, start=1):
        column_letter = get_column_letter(col_index)
        for idx, value in enumerate(df[col_name], start=2):
            cell = sheet[f"{column_letter}{idx}"]
            write_cell_value(cell, value)
            apply_cell_style(cell)

    sheet.row_dimensions[1].height = None
    return sheet
