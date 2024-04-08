from XLStyle.XLStyle import StyleReader, StyleWriter
from openpyxl import load_workbook


def GetFormatStyle(file, sheet=None, header_length=0, column_widths=[]):
    if sheet == None:
        sheet = load_workbook(file).sheetnames[0]
    FormatStyle = StyleReader(file=file, sheet=sheet, column_widths=column_widths)
    FormatStyle.GetContentStyle(header_length + 1, 1)
    FormatStyle.SetHeaderLength(header_length)
    FormatStyle.GetHeaderStyle()

    return FormatStyle


def ApplyStyle(
    style,
    file,
    write_to,
    offset_row=0,
    offset_col=0,
    applyContent=True,
    overwriteSheetSettings=False,
    wrap_columns=[],
):
    output = StyleWriter(file=file, style=style)

    if applyContent:
        output.ApplyContentStyles(
            offset_row=offset_row,
            offset_col=offset_col,
            max_row=output.Workbook[output.SheetName].max_row,
            wrap_columns=wrap_columns,
        )
    output.ApplyHeaderStyles(offset_row=offset_row, offset_col=offset_col)
    if overwriteSheetSettings:
        output.ApplySheetStyle(
            min_row=output.Style.Header_Length + 1,
            max_row=output.Workbook[output.SheetName].max_row,
        )

    output.Save(write_to)
