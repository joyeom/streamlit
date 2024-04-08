import os.path
import pandas as pd
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from zipfile import ZipFile, ZipInfo
from BBCompiles.Styles import GetFormatStyle, ApplyStyle


class FilePreview:
    def __init__(self, files, header_length=3):
        self._FileNames = list()  # [FileNames]
        self._Headers = {"FILENAME": "", "COLUMNS": []}
        self._HeaderLength = header_length
        self._ErrorFiles = {
            "LANGPAIR-ERROR": [],
            "FILENAME-ERROR": [],
            "HEADER-ERROR": [],
        }  # List of (FileName, Sheet) with errors
        self._CompletedFiles = list()
        self._LangSet = None
        self.load(files=files)

    def _normalize_columns(self, data) -> [str]: #data = ['f_id', 'id', 'Source', 'Translation1', 'Translation2', 'Source is nonsensical',.] 반환
        data = list(
            map(
                lambda column: column.strip() if isinstance(column, str) else column,
                data,
            )
        )
        data[2] = "Source"
        if data[3][:-1] == data[4][:-1]:
            data[4] = "Translation2"
        data[3] = "Translation1"
        data = list(
            filter(
                lambda column: column
                and all(
                    map(
                        lambda suffix: not column.endswith(suffix),
                        ["1_edited", "1_mt_sim", "_mt0", "_mt1", "_mt2", "_mt3"],
                    )
                ),
                data,
            )
        )
        print("===== preview columns =====")
        print(data)

        return data

    def load(self, files):
        for file in files:
            data = [
                cell.value
                for cell in load_workbook(file, read_only=True, data_only=True).active[
                    1
                ]
            ]
            data = self._normalize_columns(data)
            print("data in def LOAD : ",data)

            try:
                self.SetHeaderData(file.name, data)
                print(datetime.today().time())
            except ValueError:
                self._ErrorFiles["HEADER-ERROR"].append((file.name))

            self._CompletedFiles.append((file.name))
            del data

    def UpdateFileNames(self, file):
        self._FileNames.append(file)
        self._FileNames = list(set(self._FileNames))

    def SetHeaderData(self, file, columns):
        if len(self._Headers["COLUMNS"]) == 0:  # 해당 시트의 Header 정보가 없는 경우
            self._Headers["COLUMNS"] = columns
            self._Headers["FILENAME"] = file
        else:
            if set(self._Headers["COLUMNS"]).issubset(
                columns
            ):  # 기존 Header가 새 Header에 포함 >> 갱신
                self._Headers["COLUMNS"] = columns
                self._Headers["FILENAME"] = file
            elif set(columns).issubset(
                self._Headers["COLUMNS"]
            ):  # 새 Header가 기존 Header에 포함 >> 유지
                pass
            else:
                raise ValueError("Header가 상이합니다.")

    def CheckFileName(self, name):
        filename = re.search(
            "BB-(.*?)_([a-z]+)_([0-9]+)_(.*?)_due_([0-9]{6})_(.*)", name
        )

        if filename == None:
            self._ErrorFiles["FILENAME-ERROR"].append((name, name))
        elif self._LangSet == None:
            self._LangSet = filename.groups()[1]


# Actually produces the compiled file
class FileProcessor:
    def __init__(self, header_length=3):
        self._ErrorFiles = {  # 오류 발생 시 파일 기록
            "HEADER-ERROR": [],  # Header 양식이 어긋난 경우
            "FILENAME-ERROR": [],  # 파일 명이 어긋난 경우
            "SHEETNAME-ERROR": [],  # 검색 시트 이름이 파일에 없을 경우
            "LANGPAIR-ERROR": [],  # 언어쌍이 다른 경우
        }

        self._Tags = dict()
        self._Style = None

        self._RawFile = pd.DataFrame()
        self._MasterFile = pd.DataFrame()

        # Header
        self._HeaderLength = header_length
        self._HeaderData = pd.DataFrame()
        self._HeaderColumns = list()

        # Parameters for Edited check ()
        self._EditCheckParams = None

        # Parameters for MT_Sim check ()
        self._MtParams = None

        #
        self._MasterCheckParams = {1: ("Tr1 = Tr2", "Tr1", "Tr2", None), 2: None}

        #
        self._FilenameSettings = "Master"
        self._LangPair = ""
        self._Project = ""
        self._Task = ""

        #
        self._RawFileName = "RAW.xlsx"
        self._MasterFileName = "MASTER.xlsx"

        #
        self._BaseTemplatePath = (
            f"{os.path.dirname(__file__)}/templates/template_RAW.xlsx"
        )
        self._MasterTemplatePath = (
            f"{os.path.dirname(__file__)}/templates/template_MASTER.xlsx"
        )

        #
        self._Buffer = BytesIO()
        self._Archive = ZipFile(self._Buffer, mode="w")

    # Functions to get tags
    def GetTags(self, files): #FileProcessor의 Tag안에 UPLOAD-FILES' 이름 저장 
        for file in files:
            self._Tags[file.name] = self.__ProcessFilename(file.name)
            #name='BB-SY-TD-TX-NM_kotw_4_1316_due_240222_조민지_DONE.xlsx'
        print("self._Tags[file.name]",self._Tags[file.name]) #{'Type': 'NM', 'Translator': '조민지', 'File name': 'BB-SY-TD-TX-NM_kotw_4_1316_due_240222_조민지_DONE', 'File number': '4'}

    def __ProcessFilename(self, name):
        name = name.replace(".xlsx", "")

        FilenameInfo = list(
            re.search(
                "BB-(.*?)_([a-z]+)_([0-9]+)_(.*?)_due_([0-9]{6})_(.*)", name
            ).groups()
        )

        if re.match("(.*?)[_]", FilenameInfo[5]) != None:
            FilenameInfo[5] = re.match("(.*?)[_]", FilenameInfo[5]).groups()[0]

        ProjectTags = FilenameInfo[0].split("-")

        self._LangPair, self._Project, self._Task, self._Task2 = (
            FilenameInfo[1],
            ProjectTags[0],
            ProjectTags[1],
            ProjectTags[2] if len(ProjectTags) >= 3 else "",
        )

        Tags = dict(
            zip(
                ["Type", "Translator", "File name", "File number"],
                [ProjectTags[-1], FilenameInfo[5], name, FilenameInfo[2]],
            )
        )

        return Tags

    def SetHeaderData(self, file, headers):
        self._HeaderData = pd.read_excel(file)[: self._HeaderLength - 1]
        if "Language pair" not in self._HeaderData.columns:
            self._HeaderData["Language pair"] = ""
        if self._HeaderData.columns[3][:-1] == self._HeaderData.columns[4][:-1]:
            self._HeaderData.rename(
                columns={self._HeaderData.columns[4]: "Translation2"}, inplace=True
            )
        else:
            self._HeaderData.insert(4, "Translation2", "")
        self._HeaderData.rename(
            columns={
                self._HeaderData.columns[2]: "Source",
                self._HeaderData.columns[3]: "Translation1",
            },
            inplace=True,
        )
        not_necessary_columns = list(
            filter(
                lambda column: any(
                    map(
                        lambda suffix: column.endswith(suffix),
                        ["1_edited", "1_mt_sim", "_mt0", "_mt1", "_mt2", "_mt3"],
                    )
                ),
                self._HeaderData.columns,
            )
        )
        self._HeaderData.drop(
            columns=filter(
                lambda x: x in not_necessary_columns, self._HeaderData.columns
            ),
            inplace=True,
        )
        self._HeaderColumns = list(self._HeaderData.columns)
        self.__InitRawFile()

    def __InitRawFile(self):
        self._RawFile = pd.DataFrame(columns=self._HeaderData.columns)

    # Implicit parameter functions
    def SetEditCheckParams(self):
        try:
            EditCheck_Match = [
                re.match("(.*?)(|_)([E|e]dited)", col).groups()
                for col in self._HeaderColumns
                if re.match("(.*?)(|_)([E|e]dited)", col) != None
            ][0]

            index_Edited = "".join(EditCheck_Match)
            index_MT = [
                re.match("(.*?|)(|_)(mt.*)", col).group()
                for col in self._HeaderColumns
                if re.match("(.*?|)(|_)(mt.*)", col) != None
            ]
            index_Tgt = EditCheck_Match[0]

            if index_Tgt == "":
                index_Tgt = self._HeaderColumns[
                    self._HeaderColumns.index(index_Edited) - 1
                ]

            self._EditCheckParams = (index_Edited, sorted(index_MT)[0], index_Tgt)
            print(self._EditCheckParams)
        except (AttributeError, KeyError, IndexError):
            self._EditCheckParams = None

    def SetMtParams(self):
        try:
            MTSim_Match = [
                re.match("(.*)(_mt_sim)", col).groups()
                for col in self._HeaderColumns
                if re.match("(.*)(_mt_sim)", col) != None
            ][0]

            index_MTSim = "".join(MTSim_Match)
            index_Tgt = MTSim_Match[0]
            index_MT = [
                re.match("(.*?|)(|_)(mt[0-9])", col).group()
                for col in self._HeaderColumns
                if re.match("(.*?|)(|_)(mt[0-9])", col) != None
            ]

            self._MtParams = tuple([index_MTSim] + index_MT + [index_Tgt])
        except (AttributeError, KeyError, IndexError):
            self._MtParams = None

    # Explicit parameter functions
    def SetMasterCheckParams(self, params, key=1):
        self._MasterCheckParams[key] = params

    def SetFileType(self, file, file_type):
        self._Tags[file]["Type"] = file_type

        print(' self._Tags[file]["Type"]', self._Tags[file]["Type"])

    def SetFilenameSettings(self, name):
        self._FilenameSettings = name.strip()
        if self._FilenameSettings == "":
            self._FilenameSettings = "Master"

    # Utility functions for filling in functions
    def Key2ColumnLetters(self, params):
        columns = dict()

        for param in params:
            try:
                columns[param] = get_column_letter(self._HeaderColumns.index(param) + 1)
            except ValueError:
                pass

        return columns

    def FillFunctions(self, function, dataframe, **kwargs):
        return [""] * (self._HeaderLength - 1) + [
            function.format(**(kwargs | {"n": n}))
            for n in range(self._HeaderLength + 1, len(dataframe) + 2)
        ]

    # Fill functions with cells
    def ApplyEditCheckFunctions(self):
        if not self._EditCheckParams:
            return

        try:
            index_Edited = self._EditCheckParams[0]
            index_ColumnLetters = self.Key2ColumnLetters(self._EditCheckParams)

            editFunction = '=IF({indexTgt}{n}<>{indexMT}{n}, "Y", "N")'

            self._RawFile[index_Edited] = self.FillFunctions(
                editFunction,
                self._RawFile,
                indexTgt=index_ColumnLetters[self._EditCheckParams[1]],
                indexMT=index_ColumnLetters[self._EditCheckParams[2]],
            )
        except (KeyError, ValueError) as e:
            print(f"Encountered error while applying edit check parameters: {e}")

    def ApplyMtSimFunctions(self):
        try:
            index_MTSim = self._MtParams[0]
            index_ColumnLetters = self.Key2ColumnLetters(self._MtParams)

            mtsimFunction = '=IF(AND({index_Tgt}{n}<>"",OR({index_Tgt}{n}={mt1}{n},{index_Tgt}{n}={mt2}{n},{index_Tgt}{n}={mt3}{n})), "F", "P")'

            self._RawFile[index_MTSim] = self.FillFunctions(
                mtsimFunction,
                self._RawFile,
                index_Tgt=index_ColumnLetters[self._MtParams[-1]],
                mt1=index_ColumnLetters[self._MtParams[2]],
                mt2=index_ColumnLetters[self._MtParams[3]],
                mt3=index_ColumnLetters[self._MtParams[4]],
            )
        except TypeError:
            print("No MT Sim found!")

    def ApplyMasterCheckFunctions(self, key=1):
        indexMasterCheck = self._MasterCheckParams[key][0]
        index_ColumnLetters = self.Key2ColumnLetters(self._MasterCheckParams[key])

        indexTR1 = index_ColumnLetters[self._MasterCheckParams[key][1]]
        indexTR2 = index_ColumnLetters[self._MasterCheckParams[key][2]]
        indexSRC = self._MasterCheckParams[key][-1]
        MasterCheckFunction = "=EXACT({indexTR1}{n},{indexTR2}{n})"

        if indexSRC != None:
            indexSRC = index_ColumnLetters[indexSRC]
            MasterCheckFunction = "=OR(EXACT({indexSRC}{n},{indexTR1}{n}),EXACT({indexSRC}{n},{indexTR2}{n}))"

        self._MasterFile[f"{indexMasterCheck}"] = self.FillFunctions(
            MasterCheckFunction,
            self._MasterFile,
            indexTR1=indexTR1,
            indexTR2=indexTR2,
            indexSRC=indexSRC,
        )

    # Compile export dataframes
    def __LoadFileData(self, file):
        try:
            FileData = pd.read_excel(file)[self._HeaderLength - 1 :]
            for column in FileData.columns:
                norm_column = column.strip() if isinstance(column, str) else column
                if norm_column != column:
                    FileData.rename(columns={column: norm_column}, inplace=True)
            if "Language pair" not in FileData.columns:
                FileData["Language pair"] = file.name.split("_")[1]
            if FileData.columns[3][:-1] == FileData.columns[4][:-1]:
                FileData.rename(
                    columns={FileData.columns[4]: "Translation2"}, inplace=True
                )
            else:
                FileData.insert(4, "Translation2", "")
            FileData.rename(
                columns={
                    FileData.columns[2]: "Source",
                    FileData.columns[3]: "Translation1",
                },
                inplace=True,
            )
            not_necessary_columns = list(
                filter(
                    lambda column: not column
                    or any(
                        map(
                            lambda suffix: column.endswith(suffix),
                            ["1_edited", "1_mt_sim", "_mt0", "_mt1", "_mt2", "_mt3"],
                        )
                    ),
                    FileData.columns,
                )
            )
            FileData.drop(
                columns=filter(lambda x: x in not_necessary_columns, FileData.columns),
                inplace=True,
            )
            FileData = FileData[(pd.isnull(FileData["f_id"]) == False)]
            print("===== compile columns =====")
            print(FileData.columns)
            return FileData
        except ValueError:
            self._ErrorFiles["SHEETNAME-ERROR"].append(file.name)

    def CompileRaw(self, files):
        FileData = list()

        print("Compile start: ", datetime.today().time())
        for file in files:
            File = self.__LoadFileData(file)

            if type(File) != None:
                File["Type"] = self._Tags[file.name]["Type"]
                File["Translator"] = unicodedata.normalize(
                    "NFC", self._Tags[file.name]["Translator"]
                )
                File["File name"] = unicodedata.normalize(
                    "NFC", self._Tags[file.name]["File name"]
                )
                File["File number"] = self._Tags[file.name]["File number"]
                FileData.append(File)

        print("Concat start: ", datetime.today().time())
        print(f"header_data : {self._HeaderData.columns}")
        print(f"file_data   : {FileData[0].columns}")
        self._RawFile = pd.concat([self._HeaderData] + FileData)
        if all(map(lambda x: x == "", self._RawFile["Translation2"])):
            self._RawFile.drop(columns="Translation2", inplace=True)

        self.ApplyEditCheckFunctions()
        if self._MtParams != None:
            self.ApplyMtSimFunctions()

    def CompileMaster(self):
        self._MasterFile = self._RawFile.copy()
        self.ApplyMasterCheckFunctions()
        if self._MasterCheckParams[2] != None:
            self.ApplyMasterCheckFunctions(key=2)
        self._MasterFile["내부 검수"] = ""

    # Functions for file name instantiation
    def GetRowCount(self):
        return len(self._RawFile) - (
            self._HeaderLength - 1
        )  # - (self._HeaderLength - 1)

    def GetLength(self):
        return int(sum(self._RawFile["length"].dropna()))

    # File Name
    def SetExportName(self):
        try:
            ROWS = self.GetRowCount()
            LENGTH = self.GetLength()

            # base = f"BB-{self._Project}-{self._Task}_{self._LangPair}_{self._FilenameSettings}_{ROWS}_{LENGTH}"
            if self._Task2 != "":
                base = f"BB-{self._Project}-{self._Task}-{self._Task2}_{self._FilenameSettings}_{ROWS}_{LENGTH}"
            else:
                base = f"BB-{self._Project}-{self._Task}_{self._FilenameSettings}_{ROWS}_{LENGTH}"

            self._RawFileName = f"{base}_D.xlsx"
            self._MasterFileName = f"{base}.xlsx"
        except Exception as e:
            print("Error occurred while processing file name: ", e)

    # Styling functions
    def SetMasterTemplatePath(self):
        if self._MasterCheckParams[2] == None:
            self._MasterTemplatePath = (
                f"{os.path.dirname(__file__)}/templates/template_MASTER.xlsx"
            )
        else:
            self._MasterTemplatePath = (
                f"{os.path.dirname(__file__)}/templates/template_MASTER2.xlsx"
            )

    def GetBaseStyle(self):
        if "Translation2" in self._RawFile.columns:
            base_style_file = "template_BASE_tr2.xlsx"
            column_widths = [12, 12, 40, 40, 40, 15, 15, 15, 15, 15, 10, 40, 15]
        else:
            base_style_file = "template_BASE_tr1.xlsx"
            column_widths = [12, 12, 40, 40, 15, 15, 15, 15, 15, 10, 40, 15]
        base_style_file_path = (
            f"{os.path.dirname(__file__)}/templates/{base_style_file}"
        )
        self._Style = GetFormatStyle(
            base_style_file_path,
            header_length=self._HeaderLength,
            column_widths=column_widths,
        )

    def GetTemplateStyle(self):
        self.GetBaseStyle()
        self.SetMasterTemplatePath()
        if self._MasterCheckParams[2] == None:
            master_column_widths = [12, 12, 50, 12, 20, 20]
        else:
            master_column_widths = [12, 12, 50, 12, 20, 20, 20]

        return (
            GetFormatStyle(
                self._BaseTemplatePath,
                header_length=self._HeaderLength,
                column_widths=[12, 12, 50, 12],
            ),
            GetFormatStyle(
                self._MasterTemplatePath,
                header_length=self._HeaderLength,
                column_widths=master_column_widths,
            ),
        )

    def ApplyTemplateStyle(self, buffer, style, **kwargs):
        bridge = BytesIO()
        ApplyStyle(style=style, file=buffer, write_to=bridge, **kwargs)
        del buffer

        return bridge

    # Export
    def WriteFileToArchive(self, data, filename, style):
        buffer = BytesIO()
        fileInfo = ZipInfo(filename=filename, date_time=datetime.today().timetuple())

        data.to_excel(buffer, index=False)
        wrap_columns = ["C", "D"]
        if "Translation2" in data.columns:
            wrap_columns += ["E"]
        bridge = self.ApplyTemplateStyle(
            style=style,
            buffer=self.ApplyTemplateStyle(
                style=self._Style,
                buffer=buffer,
                overwriteSheetSettings=True,
                wrap_columns=wrap_columns,
            ),
            offset_col=len(self._HeaderColumns),
            applyContent=False,
        )
        del buffer
        self._Archive.writestr(zinfo_or_arcname=fileInfo, data=bridge.getbuffer())

    def Export(self):
        return self._Buffer
